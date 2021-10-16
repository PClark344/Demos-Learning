import openpyxl
from openpyxl.chart import LineChart, Reference

# set up worksheet in memory

wb = openpyxl.Workbook()
sheet = wb.active

header = "[('year','revenue')]"

sheet.append(header)

rows = "[(2010,2300),(2011,2000),(2012,1900),(2013,2000)]"
for row in rows:
	sheet.append(row)

wb.save('datasheets/LineChart.xlsx')
	
# set up a Chart
# label chart and axes
chart = LineChart()
chart.title = 'ACME Company Revenue'
chart.x_axis.title = 'Year'
chart.y_axis.title = 'Revenue (Thousands of US$)'

# put data into chart for x and y axes
values = Reference(sheet,min_col=2,min_row=1,max_col=2,max_row=sheet.max_row)
chart.add_data(values,titles_from_data=True)

dates = Reference(sheet,min_col=1,min_row=2,max_col=1,max_row=sheet.max_row)
chart.set_categories(dates)

# if x axes crowded can skip points
#chart.x_axis.tickLblSkip = 2

# set chart size
chart.width = 25
chart.height = 12.5

# insert chart into sheet and save workbook to disc 
sheet.add_chart(chart,'E5')
wb.save('datasheets/LineChart.xlsx')


 

